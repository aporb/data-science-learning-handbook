"""
Automated Compliance Reporting Engine
====================================

This module provides comprehensive automated reporting capabilities that generate
compliance reports in various government formats including FISMA, FedRAMP, and DoD
requirements with integration to existing monitoring and audit systems.

Key Features:
- Scheduled compliance reports (daily, weekly, monthly, quarterly)
- Government reporting format automation (FISMA, FedRAMP, DoD)
- Custom report generation with templates
- Multi-format export (PDF, Excel, JSON, XML)
- Integration with existing audit and monitoring infrastructure
- Automated report distribution and archival

Integration Points:
- Enhanced monitoring system for real-time compliance data
- Integrated audit orchestrator for comprehensive audit information
- Enhanced log aggregator for event correlation and analysis
- Multi-classification engine for classified report handling
- RBAC system for report access control

Classification: UNCLASSIFIED//FOR OFFICIAL USE ONLY
Version: 1.0 - Comprehensive Automated Reporting
Author: Security Compliance Team
Date: 2025-07-28
"""

import asyncio
import json
import logging
import time
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional, Any, Tuple, Set, Union, AsyncGenerator
from uuid import UUID, uuid4
from dataclasses import dataclass, field, asdict
from enum import Enum
from concurrent.futures import ThreadPoolExecutor
from collections import defaultdict, deque
import aiofiles
import aiohttp
from threading import Lock
import numpy as np
from pathlib import Path
import jinja2
import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Import existing infrastructure
from ...audits.integrated_audit_orchestrator import IntegratedAuditOrchestrator
from ...audits.enhanced_monitoring_system import EnhancedMonitoringSystem
from ...audits.enhanced_log_aggregator import EnhancedLogAggregator
from ...audits.audit_logger import AuditLogger, AuditEvent, AuditEventType, AuditSeverity

# Import compliance dashboard for data
from ..dashboards.compliance_dashboard import ComplianceDataProvider, ComplianceMetric, ComplianceMetricType

# Import RBAC for access control
from ...auth.rbac_system import RBACController

logger = logging.getLogger(__name__)


class ReportType(Enum):
    """Types of compliance reports."""
    FISMA_CONTINUOUS_MONITORING = "fisma_continuous_monitoring"
    FEDRAMP_ANNUAL_ASSESSMENT = "fedramp_annual_assessment"
    DOD_QUARTERLY_COMPLIANCE = "dod_quarterly_compliance"
    EXECUTIVE_SUMMARY = "executive_summary"
    TECHNICAL_ASSESSMENT = "technical_assessment"
    INCIDENT_SUMMARY = "incident_summary"
    VULNERABILITY_ASSESSMENT = "vulnerability_assessment"
    AUDIT_FINDINGS = "audit_findings"
    CONTROL_EFFECTIVENESS = "control_effectiveness"
    CUSTOM = "custom"


class ReportFormat(Enum):
    """Report output formats."""
    PDF = "pdf"
    EXCEL = "excel"
    JSON = "json"
    XML = "xml"
    HTML = "html"
    CSV = "csv"


class ReportFrequency(Enum):
    """Report generation frequency."""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    ANNUALLY = "annually"
    ON_DEMAND = "on_demand"


class ReportStatus(Enum):
    """Report generation status."""
    SCHEDULED = "scheduled"
    GENERATING = "generating"
    COMPLETED = "completed"
    FAILED = "failed"
    DELIVERED = "delivered"
    ARCHIVED = "archived"


@dataclass
class ReportConfiguration:
    """Configuration for automated report generation."""
    report_id: str = field(default_factory=lambda: str(uuid4()))
    report_type: ReportType = ReportType.EXECUTIVE_SUMMARY
    name: str = ""
    description: str = ""
    
    # Generation settings
    frequency: ReportFrequency = ReportFrequency.MONTHLY
    output_formats: List[ReportFormat] = field(default_factory=lambda: [ReportFormat.PDF])
    template_name: str = "default"
    
    # Data filters
    time_range_days: int = 30
    classification_filter: str = "UNCLASSIFIED"
    include_metrics: List[ComplianceMetricType] = field(default_factory=list)
    exclude_sensitive_data: bool = True
    
    # Access control
    required_clearance: str = "UNCLASSIFIED"  
    authorized_recipients: List[str] = field(default_factory=list)
    distribution_list: List[str] = field(default_factory=list)
    
    # Scheduling
    schedule_enabled: bool = True
    next_run: Optional[datetime] = None
    last_run: Optional[datetime] = None
    
    # Output settings
    output_directory: str = "/tmp/compliance_reports"
    filename_template: str = "{report_type}_{timestamp}"
    retention_days: int = 2555  # 7 years for DoD compliance
    
    # Metadata
    created_by: str = ""
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    last_modified: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "report_id": self.report_id,
            "report_type": self.report_type.value,
            "name": self.name,
            "description": self.description,
            "frequency": self.frequency.value,
            "output_formats": [fmt.value for fmt in self.output_formats],
            "template_name": self.template_name,
            "time_range_days": self.time_range_days,
            "classification_filter": self.classification_filter,
            "include_metrics": [metric.value for metric in self.include_metrics],
            "required_clearance": self.required_clearance,
            "authorized_recipients": self.authorized_recipients,
            "distribution_list": self.distribution_list,
            "schedule_enabled": self.schedule_enabled,
            "next_run": self.next_run.isoformat() if self.next_run else None,
            "last_run": self.last_run.isoformat() if self.last_run else None,
            "output_directory": self.output_directory,
            "retention_days": self.retention_days,
            "created_by": self.created_by,
            "created_at": self.created_at.isoformat(),
            "last_modified": self.last_modified.isoformat()
        }


@dataclass
class ReportData:
    """Structured data for report generation."""
    report_id: str
    generation_timestamp: datetime
    time_period: Tuple[datetime, datetime]
    
    # Core compliance data
    metrics: List[ComplianceMetric] = field(default_factory=list)
    summary: Dict[str, Any] = field(default_factory=dict)
    alerts: List[Dict[str, Any]] = field(default_factory=list)
    trends: Dict[str, Any] = field(default_factory=dict)
    
    # System health data
    system_health: Dict[str, Any] = field(default_factory=dict)
    integration_status: Dict[str, Any] = field(default_factory=dict)
    performance_metrics: Dict[str, Any] = field(default_factory=dict)
    
    # Audit and monitoring data
    audit_events: List[Dict[str, Any]] = field(default_factory=list)
    security_incidents: List[Dict[str, Any]] = field(default_factory=list)
    compliance_violations: List[Dict[str, Any]] = field(default_factory=list)
    
    # Government-specific data
    control_assessments: Dict[str, Any] = field(default_factory=dict)
    regulatory_status: Dict[str, Any] = field(default_factory=dict)
    recommendations: List[str] = field(default_factory=list)
    
    # Classification and access information
    classification_level: str = "UNCLASSIFIED"
    handling_instructions: str = ""
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for serialization."""
        return {
            "report_id": self.report_id,
            "generation_timestamp": self.generation_timestamp.isoformat(),
            "time_period": {
                "start": self.time_period[0].isoformat(),
                "end": self.time_period[1].isoformat()
            },
            "metrics": [metric.to_dict() for metric in self.metrics],
            "summary": self.summary,
            "alerts": self.alerts,
            "trends": self.trends,
            "system_health": self.system_health,
            "integration_status": self.integration_status,
            "performance_metrics": self.performance_metrics,
            "audit_events": self.audit_events,
            "security_incidents": self.security_incidents,
            "compliance_violations": self.compliance_violations,
            "control_assessments": self.control_assessments,
            "regulatory_status": self.regulatory_status,
            "recommendations": self.recommendations,
            "classification_level": self.classification_level,
            "handling_instructions": self.handling_instructions
        }


@dataclass
class GeneratedReport:
    """Information about a generated report."""
    report_id: str
    config_id: str
    generation_id: str = field(default_factory=lambda: str(uuid4()))
    
    # Generation metadata
    generation_timestamp: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    generation_duration_seconds: float = 0.0
    status: ReportStatus = ReportStatus.GENERATING
    
    # Report details
    report_type: ReportType = ReportType.EXECUTIVE_SUMMARY
    output_formats: List[ReportFormat] = field(default_factory=list)
    time_period: Optional[Tuple[datetime, datetime]] = None
    
    # File information
    output_files: Dict[str, str] = field(default_factory=dict)  # format -> file_path
    file_sizes: Dict[str, int] = field(default_factory=dict)    # format -> size_bytes
    
    # Data summary
    metrics_count: int = 0
    alerts_count: int = 0
    incidents_count: int = 0
    
    # Classification and access
    classification_level: str = "UNCLASSIFIED"
    authorized_for: List[str] = field(default_factory=list)
    
    # Error information
    error_message: Optional[str] = None
    warnings: List[str] = field(default_factory=list)
    
    # Delivery and archival
    delivered_to: List[str] = field(default_factory=list)
    archived_at: Optional[datetime] = None
    retention_until: Optional[datetime] = None
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "report_id": self.report_id,
            "config_id": self.config_id,
            "generation_id": self.generation_id,
            "generation_timestamp": self.generation_timestamp.isoformat(),
            "generation_duration_seconds": self.generation_duration_seconds,
            "status": self.status.value,
            "report_type": self.report_type.value,
            "output_formats": [fmt.value for fmt in self.output_formats],
            "time_period": {
                "start": self.time_period[0].isoformat(),
                "end": self.time_period[1].isoformat()
            } if self.time_period else None,
            "output_files": self.output_files,
            "file_sizes": self.file_sizes,
            "metrics_count": self.metrics_count,
            "alerts_count": self.alerts_count,
            "incidents_count": self.incidents_count,
            "classification_level": self.classification_level,
            "authorized_for": self.authorized_for,
            "error_message": self.error_message,
            "warnings": self.warnings,
            "delivered_to": self.delivered_to,
            "archived_at": self.archived_at.isoformat() if self.archived_at else None,
            "retention_until": self.retention_until.isoformat() if self.retention_until else None
        }


class ReportDataCollector:
    """Collects data from various sources for report generation."""
    
    def __init__(
        self,
        data_provider: ComplianceDataProvider,
        audit_orchestrator: IntegratedAuditOrchestrator,
        monitoring_system: EnhancedMonitoringSystem
    ):
        """Initialize report data collector."""
        self.data_provider = data_provider
        self.audit_orchestrator = audit_orchestrator
        self.monitoring_system = monitoring_system
        
        logger.info("Report data collector initialized")
    
    async def collect_report_data(
        self,
        report_config: ReportConfiguration,
        time_range: Tuple[datetime, datetime]
    ) -> ReportData:
        """Collect comprehensive data for report generation."""
        try:
            start_time = time.time()
            
            # Initialize report data
            report_data = ReportData(
                report_id=report_config.report_id,
                generation_timestamp=datetime.now(timezone.utc),
                time_period=time_range,
                classification_level=report_config.classification_filter
            )
            
            # Collect compliance metrics
            if report_config.include_metrics:
                metrics = await self.data_provider.get_compliance_metrics(
                    metric_types=report_config.include_metrics,
                    time_range=time_range,
                    classification_filter=report_config.classification_filter
                )
                report_data.metrics = metrics
            else:
                # Get all metrics if none specified
                all_metrics = await self.data_provider.get_compliance_metrics(
                    time_range=time_range,
                    classification_filter=report_config.classification_filter
                )
                report_data.metrics = all_metrics
            
            # Generate summary from metrics
            report_data.summary = self._generate_metrics_summary(report_data.metrics)
            
            # Extract alerts from metrics
            report_data.alerts = self._extract_metric_alerts(report_data.metrics)
            
            # Generate trends analysis
            report_data.trends = self._generate_trends_analysis(report_data.metrics)
            
            # Collect system health data
            report_data.system_health = await self._collect_system_health()
            
            # Collect integration status
            report_data.integration_status = self.audit_orchestrator.get_integration_status()
            
            # Collect performance metrics
            report_data.performance_metrics = self._collect_performance_metrics()
            
            # Collect audit events (sample for summary)
            report_data.audit_events = await self._collect_audit_events(time_range, limit=100)
            
            # Collect security incidents
            report_data.security_incidents = await self._collect_security_incidents(time_range)
            
            # Generate control assessments for government reports
            if report_config.report_type in [
                ReportType.FISMA_CONTINUOUS_MONITORING,
                ReportType.FEDRAMP_ANNUAL_ASSESSMENT,
                ReportType.DOD_QUARTERLY_COMPLIANCE
            ]:
                report_data.control_assessments = self._generate_control_assessments(report_data.metrics)
                report_data.regulatory_status = self._assess_regulatory_status(report_data.metrics)
            
            # Generate recommendations
            report_data.recommendations = self._generate_recommendations(report_data)
            
            # Set handling instructions based on classification
            report_data.handling_instructions = self._get_handling_instructions(
                report_config.classification_filter
            )
            
            collection_time = time.time() - start_time
            logger.info(f"Report data collected in {collection_time:.2f} seconds")
            
            return report_data
            
        except Exception as e:
            logger.error(f"Error collecting report data: {e}")
            raise
    
    def _generate_metrics_summary(self, metrics: List[ComplianceMetric]) -> Dict[str, Any]:
        """Generate summary from compliance metrics."""
        if not metrics:
            return {"overall_score": 0, "total_metrics": 0, "status": "No Data"}
        
        # Calculate overall compliance score
        scores = [metric.calculate_compliance_score() for metric in metrics]
        overall_score = sum(scores) / len(scores)
        
        # Count metrics by status
        status_counts = {"excellent": 0, "good": 0, "fair": 0, "poor": 0}
        for metric in metrics:
            score = metric.calculate_compliance_score()
            if score >= 95:
                status_counts["excellent"] += 1
            elif score >= 85:
                status_counts["good"] += 1
            elif score >= 75:
                status_counts["fair"] += 1
            else:
                status_counts["poor"] += 1
        
        # Group by framework
        framework_scores = defaultdict(list)
        for metric in metrics:
            if metric.framework:
                framework_scores[metric.framework].append(metric.calculate_compliance_score())
        
        framework_averages = {
            framework: sum(scores) / len(scores)
            for framework, scores in framework_scores.items()
        }
        
        return {
            "overall_score": round(overall_score, 1),
            "total_metrics": len(metrics),
            "status_distribution": status_counts,
            "framework_scores": framework_averages,
            "metrics_by_type": self._count_metrics_by_type(metrics)
        }
    
    def _extract_metric_alerts(self, metrics: List[ComplianceMetric]) -> List[Dict[str, Any]]:
        """Extract active alerts from metrics."""
        alerts = []
        
        for metric in metrics:
            alert_level = metric.get_alert_level()
            if alert_level.value in ["warning", "critical", "emergency"]:
                alerts.append({
                    "metric_name": metric.metric_name,
                    "alert_level": alert_level.value,
                    "message": metric.alert_message,
                    "current_value": metric.current_value,
                    "target_value": metric.target_value,
                    "framework": metric.framework,
                    "control_reference": metric.control_reference,
                    "timestamp": metric.timestamp.isoformat()
                })
        
        # Sort by severity
        severity_order = {"emergency": 0, "critical": 1, "warning": 2}
        alerts.sort(key=lambda x: severity_order.get(x["alert_level"], 3))
        
        return alerts
    
    def _generate_trends_analysis(self, metrics: List[ComplianceMetric]) -> Dict[str, Any]:
        """Generate trends analysis from metrics."""
        trends = {
            "improving": [],
            "declining": [],
            "stable": [],
            "summary": {}
        }
        
        for metric in metrics:
            trend_data = {
                "metric_name": metric.metric_name,
                "current_value": metric.current_value,
                "variance_percentage": metric.variance_percentage,
                "framework": metric.framework
            }
            trends[metric.trend_direction].append(trend_data)
        
        # Generate trend summary
        trends["summary"] = {
            "improving_count": len(trends["improving"]),
            "declining_count": len(trends["declining"]),
            "stable_count": len(trends["stable"]),
            "overall_trend": self._determine_overall_trend(trends)
        }
        
        return trends
    
    async def _collect_system_health(self) -> Dict[str, Any]:
        """Collect system health information."""
        try:
            health_check = await self.audit_orchestrator.health_check()
            
            return {
                "overall_status": health_check.get("status", "unknown"),
                "components": health_check.get("components", {}),
                "timestamp": health_check.get("timestamp"),
                "integration_health": health_check.get("integrations", {})
            }
            
        except Exception as e:
            logger.error(f"Error collecting system health: {e}")
            return {"status": "error", "error": str(e)}
    
    def _collect_performance_metrics(self) -> Dict[str, Any]:
        """Collect performance metrics from monitoring system."""
        try:
            return self.monitoring_system.get_performance_metrics()
        except Exception as e:
            logger.error(f"Error collecting performance metrics: {e}")
            return {"error": str(e)}
    
    async def _collect_audit_events(self, time_range: Tuple[datetime, datetime], limit: int = 100) -> List[Dict[str, Any]]:
        """Collect sample audit events for report."""
        try:
            # This would typically query the tamper-proof storage
            # For now, return sample structure
            return [
                {
                    "event_id": str(uuid4()),
                    "timestamp": time_range[0].isoformat(),
                    "event_type": "sample_event",
                    "severity": "medium",
                    "description": "Sample audit event for report"
                }
            ]
        except Exception as e:
            logger.error(f"Error collecting audit events: {e}")
            return []
    
    async def _collect_security_incidents(self, time_range: Tuple[datetime, datetime]) -> List[Dict[str, Any]]:
        """Collect security incidents for report."""
        try:
            # Get threat data from monitoring system
            monitoring_metrics = self.monitoring_system.get_performance_metrics()
            threat_metrics = monitoring_metrics.get("threat_detector", {})
            
            # Convert to incident format
            incidents = []
            active_threats = threat_metrics.get("active_threats", 0)
            
            if active_threats > 0:
                incidents.append({
                    "incident_id": str(uuid4()),
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "incident_type": "Active Security Threats",
                    "severity": "high" if active_threats > 5 else "medium",
                    "description": f"{active_threats} active security threats detected",
                    "status": "investigating"
                })
            
            return incidents
            
        except Exception as e:
            logger.error(f"Error collecting security incidents: {e}")
            return []
    
    def _generate_control_assessments(self, metrics: List[ComplianceMetric]) -> Dict[str, Any]:
        """Generate control effectiveness assessments for government reports."""
        assessments = {}
        
        # Group metrics by framework and control
        framework_controls = defaultdict(dict)
        
        for metric in metrics:
            if metric.framework and metric.control_reference:
                framework_controls[metric.framework][metric.control_reference] = {
                    "control_name": metric.metric_name,
                    "effectiveness_score": metric.calculate_compliance_score(),
                    "status": "effective" if metric.calculate_compliance_score() >= 85 else "needs_improvement",
                    "last_assessment": metric.last_calculation.isoformat(),
                    "findings": metric.alert_message if metric.alert_message else "No issues identified"
                }
        
        # Calculate framework-level assessments
        for framework, controls in framework_controls.items():
            scores = [control["effectiveness_score"] for control in controls.values()]
            assessments[framework] = {
                "average_effectiveness": sum(scores) / len(scores) if scores else 0,
                "total_controls": len(controls),
                "effective_controls": sum(1 for control in controls.values() 
                                        if control["status"] == "effective"),
                "controls": controls
            }
        
        return assessments
    
    def _assess_regulatory_status(self, metrics: List[ComplianceMetric]) -> Dict[str, Any]:
        """Assess overall regulatory compliance status."""
        frameworks = ["DoD 8500.01E", "NIST SP 800-53", "FISMA"]
        status = {}
        
        for framework in frameworks:
            framework_metrics = [m for m in metrics if m.framework == framework]
            if framework_metrics:
                scores = [m.calculate_compliance_score() for m in framework_metrics]
                avg_score = sum(scores) / len(scores)
                
                status[framework] = {
                    "compliance_percentage": round(avg_score, 1),
                    "status": "compliant" if avg_score >= 85 else "non_compliant",
                    "metrics_count": len(framework_metrics),
                    "last_assessment": max(m.last_calculation for m in framework_metrics).isoformat()
                }
            else:
                status[framework] = {
                    "compliance_percentage": 0,
                    "status": "not_assessed",
                    "metrics_count": 0
                }
        
        return status
    
    def _generate_recommendations(self, report_data: ReportData) -> List[str]:
        """Generate recommendations based on report data."""
        recommendations = []
        
        # Analyze alerts for recommendations
        critical_alerts = [alert for alert in report_data.alerts 
                         if alert["alert_level"] in ["critical", "emergency"]]
        
        if critical_alerts:
            recommendations.append(
                f"Address {len(critical_alerts)} critical compliance issues requiring immediate attention"
            )
        
        # Analyze declining trends
        declining_metrics = report_data.trends.get("declining", [])
        if len(declining_metrics) > 3:
            recommendations.append(
                "Review processes for metrics showing declining trends and implement corrective measures"
            )
        
        # Analyze system health
        if report_data.system_health.get("overall_status") != "healthy":
            recommendations.append(
                "Investigate and resolve system health issues affecting compliance monitoring"
            )
        
        # Framework-specific recommendations
        framework_scores = report_data.summary.get("framework_scores", {})
        for framework, score in framework_scores.items():
            if score < 85:
                recommendations.append(
                    f"Improve {framework} compliance - current score {score:.1f}% below target"
                )
        
        # Default recommendations if none generated
        if not recommendations:
            recommendations.append("Continue monitoring compliance posture and maintain current security controls")
        
        return recommendations
    
    def _get_handling_instructions(self, classification_level: str) -> str:
        """Get handling instructions based on classification level."""
        instructions = {
            "UNCLASSIFIED": "This report contains unclassified information and may be shared as appropriate.",
            "CONFIDENTIAL": "This report contains confidential information. Handle according to security protocols.",
            "SECRET": "This report contains secret information. Restrict access to authorized personnel only.",
            "TOP_SECRET": "This report contains top secret information. Handle with maximum security precautions."
        }
        
        return instructions.get(classification_level, instructions["UNCLASSIFIED"])
    
    def _count_metrics_by_type(self, metrics: List[ComplianceMetric]) -> Dict[str, int]:
        """Count metrics by type."""
        type_counts = defaultdict(int)
        for metric in metrics:
            type_counts[metric.metric_type.value] += 1
        return dict(type_counts)
    
    def _determine_overall_trend(self, trends: Dict[str, Any]) -> str:
        """Determine overall trend direction."""
        improving = len(trends["improving"])
        declining = len(trends["declining"])
        stable = len(trends["stable"])
        
        if improving > declining and improving > stable:
            return "improving"
        elif declining > improving and declining > stable:
            return "declining"
        else:
            return "stable"


class ReportGenerator:
    """Generates reports in various formats from collected data."""
    
    def __init__(self, template_directory: str = "/tmp/report_templates"):
        """Initialize report generator."""
        self.template_directory = Path(template_directory)
        self.template_directory.mkdir(parents=True, exist_ok=True)
        
        # Initialize Jinja2 template environment
        self.jinja_env = jinja2.Environment(
            loader=jinja2.FileSystemLoader(str(self.template_directory)),
            autoescape=jinja2.select_autoescape(['html', 'xml'])
        )
        
        # Create default templates
        self._create_default_templates()
        
        logger.info("Report generator initialized")
    
    def _create_default_templates(self):
        """Create default report templates."""
        # Executive summary template
        executive_template = """
# Executive Compliance Summary
**Report Period:** {{ time_period.start }} to {{ time_period.end }}
**Generated:** {{ generation_timestamp }}
**Classification:** {{ classification_level }}

## Overall Compliance Posture
- **Overall Score:** {{ summary.overall_score }}%
- **Total Metrics:** {{ summary.total_metrics }}
- **Status Distribution:**
  - Excellent (≥95%): {{ summary.status_distribution.excellent }}
  - Good (≥85%): {{ summary.status_distribution.good }}
  - Fair (≥75%): {{ summary.status_distribution.fair }}
  - Poor (<75%): {{ summary.status_distribution.poor }}

## Critical Alerts
{% if alerts %}
{% for alert in alerts[:5] %}
- **{{ alert.metric_name }}** ({{ alert.alert_level.title() }}): {{ alert.message }}
{% endfor %}
{% else %}
No critical alerts identified.
{% endif %}

## Compliance Framework Status
{% for framework, score in summary.framework_scores.items() %}
- **{{ framework }}:** {{ score }}%
{% endfor %}

## Key Recommendations
{% for recommendation in recommendations %}
- {{ recommendation }}
{% endfor %}

## Handling Instructions
{{ handling_instructions }}
"""
        
        template_path = self.template_directory / "executive_summary.md"
        with open(template_path, 'w') as f:
            f.write(executive_template)
    
    async def generate_report_formats(
        self,
        report_data: ReportData,
        config: ReportConfiguration,
        output_directory: Path
    ) -> Dict[str, str]:
        """Generate report in all requested formats."""
        output_files = {}
        
        for output_format in config.output_formats:
            try:
                if output_format == ReportFormat.PDF:
                    output_files["pdf"] = await self._generate_pdf(report_data, config, output_directory)
                elif output_format == ReportFormat.EXCEL:
                    output_files["excel"] = await self._generate_excel(report_data, config, output_directory)
                elif output_format == ReportFormat.JSON:
                    output_files["json"] = await self._generate_json(report_data, config, output_directory)
                elif output_format == ReportFormat.XML:
                    output_files["xml"] = await self._generate_xml(report_data, config, output_directory)
                elif output_format == ReportFormat.HTML:
                    output_files["html"] = await self._generate_html(report_data, config, output_directory)
                elif output_format == ReportFormat.CSV:
                    output_files["csv"] = await self._generate_csv(report_data, config, output_directory)
                
            except Exception as e:
                logger.error(f"Error generating {output_format.value} format: {e}")
        
        return output_files
    
    async def _generate_pdf(self, report_data: ReportData, config: ReportConfiguration, output_dir: Path) -> str:
        """Generate PDF report."""
        filename = f"{config.filename_template.format(report_type=config.report_type.value, timestamp=datetime.now().strftime('%Y%m%d_%H%M%S'))}.pdf"
        filepath = output_dir / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(str(filepath), pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            textColor=colors.darkblue,
            spaceAfter=20
        )
        story.append(Paragraph(f"Compliance Report: {config.name}", title_style))
        story.append(Spacer(1, 12))
        
        # Summary section
        story.append(Paragraph("Executive Summary", styles['Heading1']))
        summary_text = f"""
        <b>Report Period:</b> {report_data.time_period[0].strftime('%Y-%m-%d')} to {report_data.time_period[1].strftime('%Y-%m-%d')}<br/>
        <b>Overall Compliance Score:</b> {report_data.summary.get('overall_score', 'N/A')}%<br/>
        <b>Total Metrics:</b> {report_data.summary.get('total_metrics', 0)}<br/>
        <b>Classification:</b> {report_data.classification_level}<br/>
        """
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Metrics table
        if report_data.metrics:
            story.append(Paragraph("Compliance Metrics", styles['Heading2']))
            
            # Create table data
            table_data = [['Metric Name', 'Current Value', 'Target', 'Score', 'Status']]
            for metric in report_data.metrics[:20]:  # Limit to first 20 metrics
                score = metric.calculate_compliance_score()
                status = "✓" if score >= 85 else "⚠" if score >= 75 else "✗"
                table_data.append([
                    metric.metric_name[:30] + "..." if len(metric.metric_name) > 30 else metric.metric_name,
                    f"{metric.current_value:.1f}",
                    f"{metric.target_value:.1f}",
                    f"{score:.1f}%",
                    status
                ])
            
            # Create and style table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
        
        # Alerts section
        if report_data.alerts:
            story.append(Paragraph("Active Alerts", styles['Heading2']))
            for alert in report_data.alerts[:10]:  # First 10 alerts
                alert_text = f"<b>{alert['metric_name']}</b> ({alert['alert_level'].title()}): {alert['message']}"
                story.append(Paragraph(alert_text, styles['Normal']))
            story.append(Spacer(1, 12))
        
        # Recommendations
        if report_data.recommendations:
            story.append(Paragraph("Recommendations", styles['Heading2']))
            for i, recommendation in enumerate(report_data.recommendations, 1):
                story.append(Paragraph(f"{i}. {recommendation}", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        return str(filepath)
    
    async def _generate_excel(self, report_data: ReportData, config: ReportConfiguration, output_dir: Path) -> str:
        """Generate Excel report."""
        filename = f"{config.filename_template.format(report_type=config.report_type.value, timestamp=datetime.now().strftime('%Y%m%d_%H%M%S'))}.xlsx"
        filepath = output_dir / filename
        
        # Create workbook
        workbook = openpyxl.Workbook()
        
        # Summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Executive Summary"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Summary data
        summary_sheet['A1'] = "Compliance Report Summary"
        summary_sheet['A1'].font = Font(bold=True, size=16)
        
        summary_sheet['A3'] = "Report Period"
        summary_sheet['B3'] = f"{report_data.time_period[0].strftime('%Y-%m-%d')} to {report_data.time_period[1].strftime('%Y-%m-%d')}"
        
        summary_sheet['A4'] = "Overall Score"
        summary_sheet['B4'] = f"{report_data.summary.get('overall_score', 'N/A')}%"
        
        summary_sheet['A5'] = "Total Metrics"
        summary_sheet['B5'] = report_data.summary.get('total_metrics', 0)
        
        summary_sheet['A6'] = "Classification"
        summary_sheet['B6'] = report_data.classification_level
        
        # Metrics sheet
        if report_data.metrics:
            metrics_sheet = workbook.create_sheet("Compliance Metrics")
            
            # Headers
            headers = ['Metric Name', 'Type', 'Current Value', 'Target Value', 'Score %', 'Framework', 'Alert Level']
            for col, header in enumerate(headers, 1):
                cell = metrics_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            for row, metric in enumerate(report_data.metrics, 2):
                metrics_sheet.cell(row=row, column=1, value=metric.metric_name)
                metrics_sheet.cell(row=row, column=2, value=metric.metric_type.value)
                metrics_sheet.cell(row=row, column=3, value=metric.current_value)
                metrics_sheet.cell(row=row, column=4, value=metric.target_value)
                metrics_sheet.cell(row=row, column=5, value=metric.calculate_compliance_score())
                metrics_sheet.cell(row=row, column=6, value=metric.framework)
                metrics_sheet.cell(row=row, column=7, value=metric.get_alert_level().value)
        
        # Alerts sheet
        if report_data.alerts:
            alerts_sheet = workbook.create_sheet("Active Alerts")
            
            # Headers
            alert_headers = ['Metric Name', 'Alert Level', 'Message', 'Current Value', 'Target Value']
            for col, header in enumerate(alert_headers, 1):
                cell = alerts_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            for row, alert in enumerate(report_data.alerts, 2):
                alerts_sheet.cell(row=row, column=1, value=alert['metric_name'])
                alerts_sheet.cell(row=row, column=2, value=alert['alert_level'])
                alerts_sheet.cell(row=row, column=3, value=alert['message'])
                alerts_sheet.cell(row=row, column=4, value=alert['current_value'])
                alerts_sheet.cell(row=row, column=5, value=alert['target_value'])
        
        # Save workbook
        workbook.save(filepath)
        
        return str(filepath)
    
    async def _generate_json(self, report_data: ReportData, config: ReportConfiguration, output_dir: Path) -> str:
        """Generate JSON report."""
        filename = f"{config.filename_template.format(report_type=config.report_type.value, timestamp=datetime.now().strftime('%Y%m%d_%H%M%S'))}.json"
        filepath = output_dir / filename
        
        # Convert report data to JSON
        json_data = report_data.to_dict()
        json_data['report_config'] = config.to_dict()
        
        async with aiofiles.open(filepath, 'w') as f:
            await f.write(json.dumps(json_data, indent=2, default=str))
        
        return str(filepath)
    
    async def _generate_xml(self, report_data: ReportData, config: ReportConfiguration, output_dir: Path) -> str:
        """Generate XML report."""
        filename = f"{config.filename_template.format(report_type=config.report_type.value, timestamp=datetime.now().strftime('%Y%m%d_%H%M%S'))}.xml"
        filepath = output_dir / filename
        
        # Create XML structure
        root = ET.Element("ComplianceReport")
        root.set("classification", report_data.classification_level)
        root.set("generation_timestamp", report_data.generation_timestamp.isoformat())
        
        # Metadata
        metadata = ET.SubElement(root, "Metadata")
        ET.SubElement(metadata, "ReportId").text = report_data.report_id
        ET.SubElement(metadata, "ReportType").text = config.report_type.value
        ET.SubElement(metadata, "TimePeriodStart").text = report_data.time_period[0].isoformat()
        ET.SubElement(metadata, "TimePeriodEnd").text = report_data.time_period[1].isoformat()
        
        # Summary
        summary_elem = ET.SubElement(root, "Summary")
        summary = report_data.summary
        ET.SubElement(summary_elem, "OverallScore").text = str(summary.get('overall_score', 0))
        ET.SubElement(summary_elem, "TotalMetrics").text = str(summary.get('total_metrics', 0))
        
        # Metrics
        metrics_elem = ET.SubElement(root, "Metrics")
        for metric in report_data.metrics:
            metric_elem = ET.SubElement(metrics_elem, "Metric")
            metric_elem.set("id", metric.metric_id)
            ET.SubElement(metric_elem, "Name").text = metric.metric_name
            ET.SubElement(metric_elem, "Type").text = metric.metric_type.value
            ET.SubElement(metric_elem, "CurrentValue").text = str(metric.current_value)
            ET.SubElement(metric_elem, "TargetValue").text = str(metric.target_value)
            ET.SubElement(metric_elem, "Score").text = str(metric.calculate_compliance_score())
            ET.SubElement(metric_elem, "Framework").text = metric.framework
        
        # Write XML file
        tree = ET.ElementTree(root)
        tree.write(filepath, encoding='utf-8', xml_declaration=True)
        
        return str(filepath)
    
    async def _generate_html(self, report_data: ReportData, config: ReportConfiguration, output_dir: Path) -> str:
        """Generate HTML report."""
        filename = f"{config.filename_template.format(report_type=config.report_type.value, timestamp=datetime.now().strftime('%Y%m%d_%H%M%S'))}.html"
        filepath = output_dir / filename
        
        # HTML template
        html_template = """
<!DOCTYPE html>
<html>
<head>
    <title>Compliance Report - {{ config.name }}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .header { background-color: #366092; color: white; padding: 20px; }
        .summary { background-color: #f5f5f5; padding: 15px; margin: 20px 0; }
        .metric { border: 1px solid #ddd; padding: 10px; margin: 5px 0; }
        .alert { background-color: #fff3cd; border: 1px solid #ffeaa7; padding: 10px; margin: 5px 0; }
        .critical { background-color: #f8d7da; border: 1px solid #f5c6cb; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ config.name }}</h1>
        <p>Classification: {{ report_data.classification_level }}</p>
        <p>Generated: {{ report_data.generation_timestamp }}</p>
    </div>
    
    <div class="summary">
        <h2>Executive Summary</h2>
        <p><strong>Report Period:</strong> {{ report_data.time_period[0] }} to {{ report_data.time_period[1] }}</p>
        <p><strong>Overall Score:</strong> {{ report_data.summary.overall_score }}%</p>
        <p><strong>Total Metrics:</strong> {{ report_data.summary.total_metrics }}</p>
    </div>
    
    {% if report_data.alerts %}
    <h2>Active Alerts</h2>
    {% for alert in report_data.alerts %}
    <div class="alert {% if alert.alert_level == 'critical' %}critical{% endif %}">
        <strong>{{ alert.metric_name }}</strong> ({{ alert.alert_level|title }}):<br>
        {{ alert.message }}
    </div>
    {% endfor %}
    {% endif %}
    
    <h2>Compliance Metrics</h2>
    <table>
        <tr>
            <th>Metric Name</th>
            <th>Current Value</th>
            <th>Target</th>
            <th>Score</th>
            <th>Framework</th>
        </tr>
        {% for metric in report_data.metrics %}
        <tr>
            <td>{{ metric.metric_name }}</td>
            <td>{{ metric.current_value }}</td>
            <td>{{ metric.target_value }}</td>
            <td>{{ "%.1f"|format(metric.calculate_compliance_score()) }}%</td>
            <td>{{ metric.framework }}</td>
        </tr>
        {% endfor %}
    </table>
    
    <h2>Handling Instructions</h2>
    <p>{{ report_data.handling_instructions }}</p>
</body>
</html>
        """
        
        template = self.jinja_env.from_string(html_template)
        html_content = template.render(report_data=report_data, config=config)
        
        async with aiofiles.open(filepath, 'w') as f:
            await f.write(html_content)
        
        return str(filepath)
    
    async def _generate_csv(self, report_data: ReportData, config: ReportConfiguration, output_dir: Path) -> str:
        """Generate CSV report."""
        filename = f"{config.filename_template.format(report_type=config.report_type.value, timestamp=datetime.now().strftime('%Y%m%d_%H%M%S'))}.csv"
        filepath = output_dir / filename
        
        # Convert metrics to DataFrame
        metrics_data = []
        for metric in report_data.metrics:
            metrics_data.append({
                'Metric Name': metric.metric_name,
                'Metric Type': metric.metric_type.value,
                'Current Value': metric.current_value,
                'Target Value': metric.target_value,
                'Compliance Score': metric.calculate_compliance_score(),
                'Framework': metric.framework,
                'Control Reference': metric.control_reference,
                'Alert Level': metric.get_alert_level().value,
                'Trend Direction': metric.trend_direction,
                'Last Updated': metric.last_calculation.isoformat()
            })
        
        df = pd.DataFrame(metrics_data)
        df.to_csv(filepath, index=False)
        
        return str(filepath)


class AutomatedReportingEngine:
    """
    Main automated reporting engine that orchestrates report generation,
    scheduling, and distribution with integration to existing audit systems.
    """
    
    def __init__(
        self,
        data_provider: ComplianceDataProvider,
        audit_orchestrator: IntegratedAuditOrchestrator,
        monitoring_system: EnhancedMonitoringSystem,
        rbac_controller: Optional[RBACController] = None
    ):
        """Initialize automated reporting engine."""
        self.data_provider = data_provider
        self.audit_orchestrator = audit_orchestrator
        self.monitoring_system = monitoring_system
        self.rbac_controller = rbac_controller
        
        # Initialize components
        self.data_collector = ReportDataCollector(
            data_provider, audit_orchestrator, monitoring_system
        )
        self.report_generator = ReportGenerator()
        
        # Report management
        self.report_configs: Dict[str, ReportConfiguration] = {}
        self.generated_reports: Dict[str, GeneratedReport] = {}
        self.report_queue = asyncio.Queue()
        
        # Processing state
        self.is_running = False
        self.worker_tasks: List[asyncio.Task] = []
        
        # Create default output directory
        self.default_output_dir = Path("/tmp/compliance_reports")
        self.default_output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info("Automated reporting engine initialized")
    
    async def start(self):
        """Start the automated reporting engine."""
        if self.is_running:
            return
        
        self.is_running = True
        
        # Start worker tasks
        self.worker_tasks = [
            asyncio.create_task(self._report_generation_worker()),
            asyncio.create_task(self._schedule_monitoring_worker()),
            asyncio.create_task(self._cleanup_worker())
        ]
        
        logger.info("Automated reporting engine started")
    
    async def stop(self):
        """Stop the automated reporting engine."""
        self.is_running = False
        
        # Cancel worker tasks
        for task in self.worker_tasks:
            if not task.done():
                task.cancel()
        
        if self.worker_tasks:
            await asyncio.gather(*self.worker_tasks, return_exceptions=True)
        
        logger.info("Automated reporting engine stopped")
    
    async def create_report_config(self, config: ReportConfiguration) -> str:
        """Create a new report configuration."""
        try:
            # Validate configuration
            await self._validate_report_config(config)
            
            # Calculate next run time if scheduled
            if config.schedule_enabled:
                config.next_run = self._calculate_next_run(config.frequency)
            
            # Store configuration
            self.report_configs[config.report_id] = config
            
            logger.info(f"Created report configuration: {config.name} ({config.report_id})")
            return config.report_id
            
        except Exception as e:
            logger.error(f"Error creating report configuration: {e}")
            raise
    
    async def generate_report_on_demand(
        self,
        config_id: str,
        time_range: Optional[Tuple[datetime, datetime]] = None,
        user_id: Optional[str] = None
    ) -> str:
        """Generate a report on-demand."""
        try:
            config = self.report_configs.get(config_id)
            if not config:
                raise ValueError(f"Report configuration not found: {config_id}")
            
            # Validate user access if RBAC enabled
            if self.rbac_controller and user_id:
                access_granted = await self._validate_user_access(user_id, config)
                if not access_granted:
                    raise PermissionError("User does not have access to generate this report")
            
            # Set time range if not provided
            if not time_range:
                end_time = datetime.now(timezone.utc)
                start_time = end_time - timedelta(days=config.time_range_days)
                time_range = (start_time, end_time)
            
            # Queue report for generation
            generation_request = {
                "config_id": config_id,
                "time_range": time_range,
                "requested_by": user_id or "system",
                "priority": "high"
            }
            
            await self.report_queue.put(generation_request)
            
            # Create generated report entry
            generated_report = GeneratedReport(
                report_id=str(uuid4()),
                config_id=config_id,
                report_type=config.report_type,
                output_formats=config.output_formats,
                time_period=time_range,
                classification_level=config.classification_filter,
                authorized_for=config.authorized_recipients
            )
            
            self.generated_reports[generated_report.generation_id] = generated_report
            
            logger.info(f"Queued on-demand report generation: {generated_report.generation_id}")
            return generated_report.generation_id
            
        except Exception as e:
            logger.error(f"Error generating on-demand report: {e}")
            raise
    
    async def get_generated_report(self, generation_id: str) -> Optional[GeneratedReport]:
        """Get information about a generated report."""
        return self.generated_reports.get(generation_id)
    
    async def list_report_configs(self, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """List available report configurations."""
        configs = []
        
        for config in self.report_configs.values():
            # Check user access if RBAC enabled
            if self.rbac_controller and user_id:
                access_granted = await self._validate_user_access(user_id, config)
                if not access_granted:
                    continue
            
            configs.append(config.to_dict())
        
        return configs
    
    async def list_generated_reports(
        self,
        config_id: Optional[str] = None,
        limit: int = 50
    ) -> List[Dict[str, Any]]:
        """List generated reports."""
        reports = []
        
        for report in self.generated_reports.values():
            if config_id and report.config_id != config_id:
                continue
            
            reports.append(report.to_dict())
        
        # Sort by generation time (newest first)
        reports.sort(key=lambda x: x["generation_timestamp"], reverse=True)
        
        return reports[:limit]
    
    async def _report_generation_worker(self):
        """Worker task for processing report generation queue."""
        while self.is_running:
            try:
                # Get next report request
                generation_request = await asyncio.wait_for(
                    self.report_queue.get(), timeout=10.0
                )
                
                await self._process_report_generation(generation_request)
                
            except asyncio.TimeoutError:
                continue
            except Exception as e:
                logger.error(f"Error in report generation worker: {e}")
                await asyncio.sleep(5.0)
    
    async def _process_report_generation(self, request: Dict[str, Any]):
        """Process a single report generation request."""
        config_id = request["config_id"]
        time_range = request["time_range"]
        
        try:
            config = self.report_configs.get(config_id)
            if not config:
                logger.error(f"Report configuration not found: {config_id}")
                return
            
            # Find corresponding generated report entry
            generation_id = None
            for gen_id, gen_report in self.generated_reports.items():
                if (gen_report.config_id == config_id and 
                    gen_report.status == ReportStatus.SCHEDULED):
                    generation_id = gen_id
                    break
            
            if not generation_id:
                # Create new generated report entry for scheduled reports
                generated_report = GeneratedReport(
                    report_id=str(uuid4()),
                    config_id=config_id,
                    report_type=config.report_type,
                    output_formats=config.output_formats,
                    time_period=time_range,
                    classification_level=config.classification_filter,
                    authorized_for=config.authorized_recipients
                )
                generation_id = generated_report.generation_id
                self.generated_reports[generation_id] = generated_report
            else:
                generated_report = self.generated_reports[generation_id]
            
            # Update status
            generated_report.status = ReportStatus.GENERATING
            start_time = time.time()
            
            logger.info(f"Starting report generation: {generation_id}")
            
            # Collect report data
            report_data = await self.data_collector.collect_report_data(config, time_range)
            
            # Set up output directory
            output_directory = Path(config.output_directory)
            output_directory.mkdir(parents=True, exist_ok=True)
            
            # Generate report formats
            output_files = await self.report_generator.generate_report_formats(
                report_data, config, output_directory
            )
            
            # Update generated report with results
            generated_report.output_files = output_files
            generated_report.file_sizes = {
                fmt: Path(filepath).stat().st_size 
                for fmt, filepath in output_files.items()
            }
            generated_report.metrics_count = len(report_data.metrics)
            generated_report.alerts_count = len(report_data.alerts)
            generated_report.incidents_count = len(report_data.security_incidents)
            generated_report.generation_duration_seconds = time.time() - start_time
            generated_report.status = ReportStatus.COMPLETED
            
            # Update config last run time
            config.last_run = datetime.now(timezone.utc)
            if config.schedule_enabled:
                config.next_run = self._calculate_next_run(config.frequency)
            
            logger.info(f"Completed report generation: {generation_id} in {generated_report.generation_duration_seconds:.2f} seconds")
            
        except Exception as e:
            logger.error(f"Error processing report generation: {e}")
            
            # Update status on error
            if generation_id and generation_id in self.generated_reports:
                self.generated_reports[generation_id].status = ReportStatus.FAILED
                self.generated_reports[generation_id].error_message = str(e)
    
    async def _schedule_monitoring_worker(self):
        """Worker task for monitoring scheduled reports."""
        while self.is_running:
            try:
                current_time = datetime.now(timezone.utc)
                
                # Check for scheduled reports that need to run
                for config in self.report_configs.values():
                    if (config.schedule_enabled and 
                        config.next_run and 
                        config.next_run <= current_time):
                        
                        # Calculate time range for scheduled report
                        end_time = current_time
                        start_time = end_time - timedelta(days=config.time_range_days)
                        time_range = (start_time, end_time)
                        
                        # Queue for generation
                        generation_request = {
                            "config_id": config.report_id,
                            "time_range": time_range,
                            "requested_by": "scheduler",
                            "priority": "normal"
                        }
                        
                        await self.report_queue.put(generation_request)
                        
                        logger.info(f"Scheduled report queued: {config.name}")
                
                # Sleep for 60 seconds before next check
                await asyncio.sleep(60.0)
                
            except Exception as e:
                logger.error(f"Error in schedule monitoring worker: {e}")
                await asyncio.sleep(300.0)  # 5 minute retry delay
    
    async def _cleanup_worker(self):
        """Worker task for cleaning up old reports."""
        while self.is_running:
            try:
                current_time = datetime.now(timezone.utc)
                
                # Clean up old generated reports
                expired_reports = []
                for generation_id, report in self.generated_reports.items():
                    config = self.report_configs.get(report.config_id)
                    if config:
                        retention_period = timedelta(days=config.retention_days)
                        if current_time - report.generation_timestamp > retention_period:
                            expired_reports.append(generation_id)
                
                # Remove expired reports and files
                for generation_id in expired_reports:
                    report = self.generated_reports[generation_id]
                    
                    # Delete report files
                    for filepath in report.output_files.values():
                        try:
                            Path(filepath).unlink(missing_ok=True)
                        except Exception as e:
                            logger.warning(f"Error deleting report file {filepath}: {e}")
                    
                    # Remove from registry
                    del self.generated_reports[generation_id]
                    logger.info(f"Cleaned up expired report: {generation_id}")
                
                # Run cleanup daily
                await asyncio.sleep(86400.0)  # 24 hours
                
            except Exception as e:
                logger.error(f"Error in cleanup worker: {e}")
                await asyncio.sleep(3600.0)  # 1 hour retry delay
    
    async def _validate_report_config(self, config: ReportConfiguration):
        """Validate report configuration."""
        if not config.name:
            raise ValueError("Report name is required")
        
        if not config.output_formats:
            raise ValueError("At least one output format is required")
        
        if config.time_range_days <= 0:
            raise ValueError("Time range days must be positive")
        
        # Validate output directory
        output_dir = Path(config.output_directory)
        if not output_dir.exists():
            try:
                output_dir.mkdir(parents=True)
            except Exception as e:
                raise ValueError(f"Cannot create output directory: {e}")
    
    async def _validate_user_access(self, user_id: str, config: ReportConfiguration) -> bool:
        """Validate user access to report configuration."""
        try:
            if not self.rbac_controller:
                return True
            
            # Check if user is in authorized recipients
            if config.authorized_recipients and user_id not in config.authorized_recipients:
                return False
            
            # Check clearance level
            # This would typically integrate with user clearance verification
            # For now, assume basic validation
            return True
            
        except Exception as e:
            logger.error(f"Error validating user access: {e}")
            return False
    
    def _calculate_next_run(self, frequency: ReportFrequency) -> datetime:
        """Calculate next run time based on frequency."""
        current_time = datetime.now(timezone.utc)
        
        if frequency == ReportFrequency.DAILY:
            return current_time + timedelta(days=1)
        elif frequency == ReportFrequency.WEEKLY:
            return current_time + timedelta(weeks=1)
        elif frequency == ReportFrequency.MONTHLY:
            return current_time + timedelta(days=30)  # Approximate
        elif frequency == ReportFrequency.QUARTERLY:
            return current_time + timedelta(days=90)  # Approximate
        elif frequency == ReportFrequency.ANNUALLY:
            return current_time + timedelta(days=365)  # Approximate
        else:
            return current_time + timedelta(days=1)  # Default to daily
    
    async def health_check(self) -> Dict[str, Any]:
        """Perform health check of reporting engine."""
        return {
            "status": "healthy" if self.is_running else "stopped",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "report_configs": len(self.report_configs),
            "generated_reports": len(self.generated_reports),
            "queue_size": self.report_queue.qsize(),
            "worker_tasks": len([t for t in self.worker_tasks if not t.done()])
        }


# Reporting manager class for easy initialization
class ReportingManager:
    """Manager class for easy initialization of the reporting system."""
    
    def __init__(
        self,
        audit_orchestrator: IntegratedAuditOrchestrator,
        monitoring_system: EnhancedMonitoringSystem,
        log_aggregator: EnhancedLogAggregator,
        classification_engine: Optional[Any] = None,
        rbac_controller: Optional[RBACController] = None
    ):
        """Initialize reporting manager."""
        # Create data provider
        self.data_provider = ComplianceDataProvider(
            audit_orchestrator, monitoring_system, log_aggregator, classification_engine
        )
        
        # Create reporting engine
        self.reporting_engine = AutomatedReportingEngine(
            self.data_provider,
            audit_orchestrator,
            monitoring_system,
            rbac_controller
        )
        
        logger.info("Reporting manager initialized")
    
    async def start(self):
        """Start the reporting system."""
        await self.reporting_engine.start()
        logger.info("Reporting system started")
    
    async def stop(self):
        """Stop the reporting system."""
        await self.reporting_engine.stop()
        logger.info("Reporting system stopped")
    
    def get_reporting_engine(self) -> AutomatedReportingEngine:
        """Get the reporting engine instance."""
        return self.reporting_engine


# Factory function
def create_reporting_manager(
    audit_orchestrator: IntegratedAuditOrchestrator,
    monitoring_system: EnhancedMonitoringSystem,
    log_aggregator: EnhancedLogAggregator,
    classification_engine: Optional[Any] = None,
    rbac_controller: Optional[RBACController] = None
) -> ReportingManager:
    """Create and initialize reporting manager."""
    return ReportingManager(
        audit_orchestrator=audit_orchestrator,
        monitoring_system=monitoring_system,
        log_aggregator=log_aggregator,
        classification_engine=classification_engine,
        rbac_controller=rbac_controller
    )


if __name__ == "__main__":
    # Example usage
    print("Automated Compliance Reporting Engine - see code for usage examples")